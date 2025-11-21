import streamlit as st
import pandas as pd
import plotly.graph_objects as go
import plotly.express as px
from budget_forecast import BudgetForecaster
import numpy as np
import tempfile
import os

# Sayfa konfigürasyonu
st.set_page_config(
    page_title="2026 Satış Bütçe Tahmini",
    page_icon="📊",
    layout="wide"
)

# CSS
st.markdown("""
    <style>
    .main-header {
        font-size: 2.5rem;
        font-weight: bold;
        color: #1f77b4;
        text-align: center;
        margin-bottom: 2rem;
    }
    </style>
""", unsafe_allow_html=True)

# Header
st.markdown('<p class="main-header">📊 2026 Satış Bütçe Tahmini Sistemi</p>', unsafe_allow_html=True)

# Sidebar başlık
st.sidebar.header("📋 Tahmin Parametreleri")

# 1. FILE UPLOAD - EN ÖNCE
st.sidebar.markdown("---")
st.sidebar.subheader("📂 Veri Yükleme")
uploaded_file = st.sidebar.file_uploader(
    "Excel Dosyası Yükle",
    type=['xlsx'],
    help="2024-2025 verilerini içeren Excel dosyası"
)

# Veri yükleme
@st.cache_data
def load_data(file_path):
    return BudgetForecaster(file_path)

forecaster = None
if uploaded_file is not None:
    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp_file:
        tmp_file.write(uploaded_file.getvalue())
        tmp_path = tmp_file.name
    
    with st.spinner('Veri yükleniyor...'):
        forecaster = load_data(tmp_path)
    
    os.unlink(tmp_path)

# Eğer dosya yüklenmemişse bilgi göster ve dur
if forecaster is None:
    st.info("👆 Lütfen soldaki menüden Excel dosyanızı yükleyin.")
    st.markdown("""
    ### Nasıl Kullanılır?
    1. Sol taraftaki **"📂 Veri Yükleme"** bölümünden Excel dosyanızı yükleyin
    2. **Büyüme hedeflerinizi** belirleyin (ay bazında ve/veya ana grup bazında)
    3. **Karlılık ve stok hedeflerinizi** ayarlayın
    4. Sistem otomatik olarak 2026 tahminini yapacak
    """)
    st.stop()

# Dosya yüklendiyse parametreleri göster
st.sidebar.markdown("---")
st.sidebar.subheader("💰 Büyüme Hedefi")

# 2. AY BAZINDA HEDEF
st.sidebar.markdown("### 📅 Ay Bazında Hedef")
monthly_input_type = st.sidebar.radio(
    "Ay Hedefi",
    ["Tüm Aylar İçin Tek Hedef", "Her Ay Ayrı Hedef"],
    index=0,
    key="monthly_type"
)

monthly_growth_targets = {}

if monthly_input_type == "Tüm Aylar İçin Tek Hedef":
    monthly_default = st.sidebar.slider(
        "Tüm Aylar İçin Büyüme Hedefi (%)",
        min_value=-20.0,
        max_value=50.0,
        value=15.0,
        step=1.0,
        key="monthly_default"
    ) / 100
    
    for month in range(1, 13):
        monthly_growth_targets[month] = monthly_default
else:
    st.sidebar.caption("↓ Aşağı kaydırarak tüm ayları görebilirsiniz")
    
    month_names = {
        1: "Ocak", 2: "Şubat", 3: "Mart", 4: "Nisan",
        5: "Mayıs", 6: "Haziran", 7: "Temmuz", 8: "Ağustos",
        9: "Eylül", 10: "Ekim", 11: "Kasım", 12: "Aralık"
    }
    
    for month in range(1, 13):
        monthly_growth_targets[month] = st.sidebar.slider(
            f"{month_names[month]} ({month})",
            min_value=-20.0,
            max_value=50.0,
            value=15.0,
            step=1.0,
            key=f"month_{month}"
        ) / 100
    
    avg_monthly = sum(monthly_growth_targets.values()) / 12
    st.sidebar.info(f"📊 Ort. Aylık: %{avg_monthly*100:.1f}")

# 3. ANA GRUP BAZINDA HEDEF
st.sidebar.markdown("---")
st.sidebar.markdown("### 🏪 Ana Grup Bazında Hedef")

# Ana grupları al (cache yok, her seferinde hesaplansın - hızlı zaten)
main_groups = sorted(forecaster.data['MainGroup'].unique().tolist())

maingroup_input_type = st.sidebar.radio(
    "Ana Grup Hedefi",
    ["Tüm Gruplar İçin Tek Hedef", "Her Grup Ayrı Hedef"],
    index=0,
    key="maingroup_type"
)

maingroup_growth_targets = {}

if maingroup_input_type == "Tüm Gruplar İçin Tek Hedef":
    maingroup_default = st.sidebar.slider(
        "Tüm Gruplar İçin Büyüme Hedefi (%)",
        min_value=-20.0,
        max_value=50.0,
        value=15.0,
        step=1.0,
        key="maingroup_default"
    ) / 100
    
    for group in main_groups:
        maingroup_growth_targets[group] = maingroup_default
else:
    st.sidebar.caption("↓ Aşağı kaydırarak tüm grupları görebilirsiniz")
    
    for group in main_groups:
        maingroup_growth_targets[group] = st.sidebar.slider(
            f"{group}",
            min_value=-20.0,
            max_value=50.0,
            value=15.0,
            step=1.0,
            key=f"group_{group}"
        ) / 100
    
    avg_maingroup = sum(maingroup_growth_targets.values()) / len(maingroup_growth_targets)
    st.sidebar.info(f"📊 Ort. Ana Grup: %{avg_maingroup*100:.1f}")

growth_param = sum(monthly_growth_targets.values()) / 12

# 4. KARLILIK HEDEFİ
st.sidebar.markdown("---")
st.sidebar.subheader("📈 Karlılık Hedefi")
margin_improvement = st.sidebar.slider(
    "Brüt Marj İyileşme Hedefi (puan)",
    min_value=-5.0,
    max_value=10.0,
    value=2.0,
    step=0.5,
    help="Mevcut brüt marj üzerine eklenecek puan"
) / 100

# 5. STOK HEDEFİ
st.sidebar.markdown("---")
st.sidebar.subheader("📦 Stok Hedefi")

stock_param_type = st.sidebar.radio(
    "Stok Parametresi",
    ["Stok/SMM Oranı", "Stok Tutar Değişimi"],
    index=0,
    help="Stok hedefini oran veya tutar bazında belirle"
)

if stock_param_type == "Stok/SMM Oranı":
    stock_ratio_target = st.sidebar.slider(
        "Hedef Stok/SMM Oranı",
        min_value=0.3,
        max_value=2.0,
        value=0.8,
        step=0.1,
        help="Stok tutarı / Satılan Malın Maliyeti oranı"
    )
    stock_change_pct = None
else:
    stock_change_pct = st.sidebar.slider(
        "Stok Tutar Değişimi (%)",
        min_value=-50.0,
        max_value=100.0,
        value=0.0,
        step=5.0,
        help="2025'e göre stok tutarında % artış veya azalış"
    ) / 100
    stock_ratio_target = None

# TAHMİN YAP
with st.spinner('Tahmin hesaplanıyor...'):
    # Stock parametresini belirle
    if stock_change_pct is not None:
        # Tutar bazlı değişim - forecaster'a direkt geç
        full_data = forecaster.get_full_data_with_forecast(
            growth_param=growth_param,
            margin_improvement=margin_improvement,
            stock_ratio_target=None,
            stock_change_pct=stock_change_pct,
            monthly_growth_targets=monthly_growth_targets,
            maingroup_growth_targets=maingroup_growth_targets
        )
    else:
        # Oran bazlı hedef - eski yöntem
        full_data = forecaster.get_full_data_with_forecast(
            growth_param=growth_param,
            margin_improvement=margin_improvement,
            stock_ratio_target=stock_ratio_target,
            stock_change_pct=None,
            monthly_growth_targets=monthly_growth_targets,
            maingroup_growth_targets=maingroup_growth_targets
        )
    
    summary = forecaster.get_summary_stats(full_data)
    quality_metrics = forecaster.get_forecast_quality_metrics(full_data)

# ANA METRİKLER
st.markdown("## 📈 Özet Metrikler")

# İLK SATIR - Ana Metrikler
col1, col2, col3, col4 = st.columns(4)

with col1:
    sales_2026 = summary[2026]['Total_Sales']
    sales_2025 = summary[2025]['Total_Sales']
    sales_growth = ((sales_2026 - sales_2025) / sales_2025 * 100) if sales_2025 > 0 else 0
    
    st.metric(
        label="2026 Toplam Satış",
        value=f"₺{sales_2026:,.0f}",
        delta=f"%{sales_growth:.1f} vs 2025"
    )

with col2:
    margin_2026 = summary[2026]['Avg_GrossMargin%']
    margin_2025 = summary[2025]['Avg_GrossMargin%']
    margin_change = margin_2026 - margin_2025
    
    st.metric(
        label="2026 Brüt Marj",
        value=f"%{margin_2026:.1f}",
        delta=f"{margin_change:+.1f} puan"
    )

with col3:
    gp_2026 = summary[2026]['Total_GrossProfit']
    gp_2025 = summary[2025]['Total_GrossProfit']
    gp_growth = ((gp_2026 - gp_2025) / gp_2025 * 100) if gp_2025 > 0 else 0
    
    st.metric(
        label="2026 Brüt Kar",
        value=f"₺{gp_2026:,.0f}",
        delta=f"%{gp_growth:.1f} vs 2025"
    )

with col4:
    if stock_change_pct is not None:
        # Stok tutar değişimi göster
        stock_2026 = summary[2026]['Avg_Stock']
        stock_2025 = summary[2025]['Avg_Stock']
        stock_change = ((stock_2026 - stock_2025) / stock_2025 * 100) if stock_2025 > 0 else 0
        
        st.metric(
            label="2026 Ort. Stok",
            value=f"₺{stock_2026:,.0f}",
            delta=f"%{stock_change:+.1f} vs 2025"
        )
    else:
        # Haftalık Stok/SMM oranı göster
        stock_weekly_2026 = summary[2026]['Avg_Stock_COGS_Weekly']
        stock_weekly_2025 = summary[2025]['Avg_Stock_COGS_Weekly']
        weekly_change = stock_weekly_2026 - stock_weekly_2025
        
        st.metric(
            label="2026 Stok/SMM (Haftalık)",
            value=f"{stock_weekly_2026:.2f} hafta",
            delta=f"{weekly_change:+.2f} hafta vs 2025"
        )
        st.caption("Stok / (Aylık SMM ÷ gün × 7)")

# İKİNCİ SATIR - Tahmin Kalite Metrikleri (Sadece Göstergeler)
st.markdown("### 🎯 Tahmin Güvenilirlik Göstergeleri")

col1, col2, col3, col4 = st.columns(4)

with col1:
    if quality_metrics['r2_score'] is not None:
        r2_pct = quality_metrics['r2_score'] * 100
        
        # Gösterge belirleme
        if r2_pct > 80:
            indicator = "🟢 Çok İyi"
        elif r2_pct > 60:
            indicator = "🟡 İyi"
        elif r2_pct > 40:
            indicator = "🟠 Orta"
        else:
            indicator = "🔴 Zayıf"
        
        st.metric(
            label="Model Uyumu",
            value=indicator,
            help="2024-2025 trend tutarlılığı"
        )
    else:
        st.metric(label="Model Uyumu", value="⚪ Hesaplanamadı")

with col2:
    if quality_metrics['trend_consistency'] is not None:
        consistency_pct = quality_metrics['trend_consistency'] * 100
        
        if consistency_pct > 80:
            indicator = "🟢 Çok İstikrarlı"
        elif consistency_pct > 60:
            indicator = "🟡 İstikrarlı"
        elif consistency_pct > 40:
            indicator = "🟠 Değişken"
        else:
            indicator = "🔴 Çok Değişken"
        
        st.metric(
            label="Trend İstikrarı",
            value=indicator,
            help="Aylık büyüme oranlarının tutarlılığı"
        )
    else:
        st.metric(label="Trend İstikrarı", value="⚪ Hesaplanamadı")

with col3:
    if quality_metrics['mape'] is not None:
        mape = quality_metrics['mape']
        
        if mape < 15:
            indicator = "🟢 Düşük Hata"
        elif mape < 25:
            indicator = "🟡 Kabul Edilebilir"
        elif mape < 35:
            indicator = "🟠 Yüksek Hata"
        else:
            indicator = "🔴 Çok Yüksek Hata"
        
        st.metric(
            label="Tahmin Hatası",
            value=indicator,
            help="Ortalama sapma oranı"
        )
    else:
        st.metric(label="Tahmin Hatası", value="⚪ Hesaplanamadı")

with col4:
    confidence = quality_metrics['confidence_level']
    
    # Genel değerlendirme
    if confidence == 'Yüksek':
        overall = "🟢 Güvenilir"
    elif confidence == 'Orta':
        overall = "🟡 Makul"
    else:
        overall = "🟠 Dikkatli Kullan"
    
    st.metric(
        label="Genel Değerlendirme",
        value=overall,
        help="Tüm metriklerin ortalaması"
    )
    
    # Organik büyümeyi göster (bu pozitif bir bilgi)
    if quality_metrics['avg_growth_2024_2025']:
        st.caption(f"📈 2024→2025 Büyüme: %{quality_metrics['avg_growth_2024_2025']:.1f}")

st.markdown("---")

# TABLAR
tab1, tab2, tab3, tab4 = st.tabs(["📊 Aylık Trend", "🎯 Ana Grup Analizi", "📅 Yıllık Karşılaştırma", "📋 Detay Veriler"])

with tab1:
    st.subheader("Aylık Satış Trendi (2024-2026)")
    
    monthly_sales = full_data.groupby(['Year', 'Month'])['Sales'].sum().reset_index()
    
    fig = go.Figure()
    
    for year in [2024, 2025, 2026]:
        year_data = monthly_sales[monthly_sales['Year'] == year]
        
        line_style = 'solid' if year < 2026 else 'dash'
        line_width = 2 if year < 2026 else 3
        
        fig.add_trace(go.Scatter(
            x=year_data['Month'],
            y=year_data['Sales'],
            mode='lines+markers',
            name=f'{year}' + (' (Tahmin)' if year == 2026 else ''),
            line=dict(dash=line_style, width=line_width),
            marker=dict(size=8)
        ))
    
    fig.update_layout(
        title="Aylık Satış Karşılaştırması",
        xaxis_title="Ay",
        yaxis_title="Satış (TRY)",
        hovermode='x unified',
        height=500
    )
    
    st.plotly_chart(fig, use_container_width=True)
    
    # Brüt Marj Trendi
    st.subheader("Aylık Brüt Marj % Trendi")
    
    monthly_margin = full_data.groupby(['Year', 'Month']).apply(
        lambda x: (x['GrossProfit'].sum() / x['Sales'].sum() * 100) if x['Sales'].sum() > 0 else 0
    ).reset_index(name='Margin%')
    
    fig2 = go.Figure()
    
    for year in [2024, 2025, 2026]:
        year_data = monthly_margin[monthly_margin['Year'] == year]
        
        line_style = 'solid' if year < 2026 else 'dash'
        
        fig2.add_trace(go.Scatter(
            x=year_data['Month'],
            y=year_data['Margin%'],
            mode='lines+markers',
            name=f'{year}' + (' (Tahmin)' if year == 2026 else ''),
            line=dict(dash=line_style),
            marker=dict(size=8)
        ))
    
    fig2.update_layout(
        title="Aylık Brüt Marj % Karşılaştırması",
        xaxis_title="Ay",
        yaxis_title="Brüt Marj %",
        hovermode='x unified',
        height=500
    )
    
    st.plotly_chart(fig2, use_container_width=True)

with tab2:
    st.subheader("Ana Grup Bazında Performans")
    
    group_sales = full_data.groupby(['Year', 'MainGroup'])['Sales'].sum().reset_index()
    
    top_groups_2026 = group_sales[group_sales['Year'] == 2026].nlargest(10, 'Sales')['MainGroup'].tolist()
    
    group_sales_filtered = group_sales[group_sales['MainGroup'].isin(top_groups_2026)]
    
    fig3 = px.bar(
        group_sales_filtered,
        x='MainGroup',
        y='Sales',
        color='Year',
        barmode='group',
        title='Top 10 Ana Grup - Yıllık Satış Karşılaştırması'
    )
    
    fig3.update_layout(height=500, xaxis_tickangle=-45)
    st.plotly_chart(fig3, use_container_width=True)
    
    # Büyüme analizi
    st.subheader("Ana Grup Büyüme Analizi (2025 → 2026)")
    
    sales_2025 = group_sales[group_sales['Year'] == 2025][['MainGroup', 'Sales']]
    sales_2025.columns = ['MainGroup', 'Sales_2025']
    
    sales_2026_grp = group_sales[group_sales['Year'] == 2026][['MainGroup', 'Sales']]
    sales_2026_grp.columns = ['MainGroup', 'Sales_2026']
    
    growth_analysis = sales_2025.merge(sales_2026_grp, on='MainGroup')
    growth_analysis['Growth%'] = ((growth_analysis['Sales_2026'] - growth_analysis['Sales_2025']) / 
                                   growth_analysis['Sales_2025'] * 100)
    growth_analysis = growth_analysis.sort_values('Growth%', ascending=False)
    
    fig4 = px.bar(
        growth_analysis.head(15),
        x='MainGroup',
        y='Growth%',
        title='Top 15 Ana Grup - Büyüme Oranı',
        color='Growth%',
        color_continuous_scale='RdYlGn'
    )
    
    fig4.update_layout(height=500, xaxis_tickangle=-45)
    st.plotly_chart(fig4, use_container_width=True)

with tab3:
    st.subheader("Yıllık Toplam Karşılaştırma")
    
    col1, col2 = st.columns(2)
    
    with col1:
        yearly_summary = pd.DataFrame({
            'Yıl': [2024, 2025, 2026],
            'Satış': [summary[2024]['Total_Sales'], 
                     summary[2025]['Total_Sales'],
                     summary[2026]['Total_Sales']],
            'Brüt Kar': [summary[2024]['Total_GrossProfit'],
                        summary[2025]['Total_GrossProfit'],
                        summary[2026]['Total_GrossProfit']]
        })
        
        fig5 = go.Figure()
        fig5.add_trace(go.Bar(name='Satış', x=yearly_summary['Yıl'], y=yearly_summary['Satış']))
        fig5.add_trace(go.Bar(name='Brüt Kar', x=yearly_summary['Yıl'], y=yearly_summary['Brüt Kar']))
        
        fig5.update_layout(
            title='Yıllık Satış ve Brüt Kar',
            barmode='group',
            height=400
        )
        
        st.plotly_chart(fig5, use_container_width=True)
    
    with col2:
        yearly_margin = pd.DataFrame({
            'Yıl': [2024, 2025, 2026],
            'Brüt Marj %': [summary[2024]['Avg_GrossMargin%'],
                           summary[2025]['Avg_GrossMargin%'],
                           summary[2026]['Avg_GrossMargin%']]
        })
        
        fig6 = go.Figure()
        fig6.add_trace(go.Scatter(
            x=yearly_margin['Yıl'],
            y=yearly_margin['Brüt Marj %'],
            mode='lines+markers',
            line=dict(width=3),
            marker=dict(size=12)
        ))
        
        fig6.update_layout(
            title='Yıllık Brüt Marj %',
            height=400,
            yaxis_title='Brüt Marj %'
        )
        
        st.plotly_chart(fig6, use_container_width=True)
    
    st.subheader("Yıllık Özet Tablo")
    
    summary_table = pd.DataFrame({
        'Metrik': ['Toplam Satış (TRY)', 'Toplam Brüt Kar (TRY)', 
                  'Brüt Marj %', 'Ort. Stok (TRY)', 'Stok/SMM Oranı'],
        '2024': [
            f"₺{summary[2024]['Total_Sales']:,.0f}",
            f"₺{summary[2024]['Total_GrossProfit']:,.0f}",
            f"%{summary[2024]['Avg_GrossMargin%']:.2f}",
            f"₺{summary[2024]['Avg_Stock']:,.0f}",
            f"{summary[2024]['Avg_Stock_COGS_Ratio']:.2f}"
        ],
        '2025': [
            f"₺{summary[2025]['Total_Sales']:,.0f}",
            f"₺{summary[2025]['Total_GrossProfit']:,.0f}",
            f"%{summary[2025]['Avg_GrossMargin%']:.2f}",
            f"₺{summary[2025]['Avg_Stock']:,.0f}",
            f"{summary[2025]['Avg_Stock_COGS_Ratio']:.2f}"
        ],
        '2026 (Tahmin)': [
            f"₺{summary[2026]['Total_Sales']:,.0f}",
            f"₺{summary[2026]['Total_GrossProfit']:,.0f}",
            f"%{summary[2026]['Avg_GrossMargin%']:.2f}",
            f"₺{summary[2026]['Avg_Stock']:,.0f}",
            f"{summary[2026]['Avg_Stock_COGS_Ratio']:.2f}"
        ]
    })
    
    st.dataframe(summary_table, use_container_width=True, hide_index=True)

with tab4:
    st.subheader("Detaylı Veri Tablosu - Yan Yana Karşılaştırma")
    
    # Ay seçimi
    selected_month = st.selectbox("Ay Seçin", list(range(1, 13)), format_func=lambda x: f"{x}. Ay")
    
    # Her yıl için veri al
    data_2024 = full_data[(full_data['Year'] == 2024) & (full_data['Month'] == selected_month)].copy()
    data_2025 = full_data[(full_data['Year'] == 2025) & (full_data['Month'] == selected_month)].copy()
    data_2026 = full_data[(full_data['Year'] == 2026) & (full_data['Month'] == selected_month)].copy()
    
    # Aylık gün sayıları
    days_in_month = {1: 31, 2: 28, 3: 31, 4: 30, 5: 31, 6: 30,
                     7: 31, 8: 31, 9: 30, 10: 31, 11: 30, 12: 31}
    days = days_in_month[selected_month]
    
    # MainGroup bazında birleştir
    comparison = data_2024[['MainGroup', 'Sales', 'GrossMargin%', 'Stock', 'COGS']].rename(
        columns={
            'Sales': 'Satış_2024',
            'GrossMargin%': 'BM%_2024',
            'Stock': 'Stok_2024',
            'COGS': 'SMM_2024'
        }
    )
    
    comparison = comparison.merge(
        data_2025[['MainGroup', 'Sales', 'GrossMargin%', 'Stock', 'COGS']].rename(
            columns={
                'Sales': 'Satış_2025',
                'GrossMargin%': 'BM%_2025',
                'Stock': 'Stok_2025',
                'COGS': 'SMM_2025'
            }
        ),
        on='MainGroup',
        how='outer'
    )
    
    comparison = comparison.merge(
        data_2026[['MainGroup', 'Sales', 'GrossMargin%', 'Stock', 'COGS']].rename(
            columns={
                'Sales': 'Satış_2026',
                'GrossMargin%': 'BM%_2026',
                'Stock': 'Stok_2026',
                'COGS': 'SMM_2026'
            }
        ),
        on='MainGroup',
        how='outer'
    )
    
    comparison = comparison.fillna(0)
    
    # Haftalık normalize - Stok/SMM Haftalık
    comparison['Stok/SMM_Haftalık_2024'] = np.where(
        comparison['SMM_2024'] > 0,
        comparison['Stok_2024'] / ((comparison['SMM_2024'] / days) * 7),
        0
    )
    comparison['Stok/SMM_Haftalık_2025'] = np.where(
        comparison['SMM_2025'] > 0,
        comparison['Stok_2025'] / ((comparison['SMM_2025'] / days) * 7),
        0
    )
    comparison['Stok/SMM_Haftalık_2026'] = np.where(
        comparison['SMM_2026'] > 0,
        comparison['Stok_2026'] / ((comparison['SMM_2026'] / days) * 7),
        0
    )
    
    # Formatla - Gösterim için
    display_df = comparison.copy()
    
    # Para formatı
    for col in ['Satış_2024', 'Stok_2024', 'SMM_2024', 'Satış_2025', 'Stok_2025', 'SMM_2025', 
                'Satış_2026', 'Stok_2026', 'SMM_2026']:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda x: f"₺{x:,.0f}" if x > 0 else "-")
    
    # Yüzde formatı
    for col in ['BM%_2024', 'BM%_2025', 'BM%_2026']:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda x: f"%{x*100:.1f}" if x > 0 else "-")
    
    # Stok/SMM Haftalık formatı
    for col in ['Stok/SMM_Haftalık_2024', 'Stok/SMM_Haftalık_2025', 'Stok/SMM_Haftalık_2026']:
        if col in display_df.columns:
            display_df[col] = display_df[col].apply(lambda x: f"{x:.2f}" if x > 0 else "-")
    
    # Sütun sırası - Yan yana karşılaştırma
    display_df = display_df[[
        'MainGroup',
        'Satış_2024', 'Satış_2025', 'Satış_2026',
        'BM%_2024', 'BM%_2025', 'BM%_2026',
        'Stok_2024', 'Stok_2025', 'Stok_2026',
        'SMM_2024', 'SMM_2025', 'SMM_2026',
        'Stok/SMM_Haftalık_2024', 'Stok/SMM_Haftalık_2025', 'Stok/SMM_Haftalık_2026'
    ]]
    
    # Sütun isimlerini güzelleştir
    display_df.columns = [
        'Ana Grup',
        'Satış 2024', 'Satış 2025', 'Satış 2026',
        'BM% 2024', 'BM% 2025', 'BM% 2026',
        'Stok 2024', 'Stok 2025', 'Stok 2026',
        'SMM 2024', 'SMM 2025', 'SMM 2026',
        'Stok/SMM Hft. 2024', 'Stok/SMM Hft. 2025', 'Stok/SMM Hft. 2026'
    ]
    
    st.info(f"📅 {selected_month}. Ay ({days} gün) - Stok/SMM haftalık: (Stok / (SMM/{days})*7)")
    
    st.dataframe(
        display_df,
        use_container_width=True,
        hide_index=True,
        height=600
    )
    
    # Excel export - ham veri
    st.download_button(
        label="📥 CSV İndir (Sadece Bu Ay)",
        data=comparison.to_csv(index=False).encode('utf-8'),
        file_name=f'budget_comparison_month_{selected_month}.csv',
        mime='text/csv'
    )
    
    # Tam Excel dosyası oluştur
    st.markdown("---")
    st.subheader("📊 Tam Bütçe Dosyası İndir")
    
    if st.button("🔄 Excel Dosyası Oluştur (Orijinal + 2026)", type="primary"):
        with st.spinner("Excel dosyası hazırlanıyor..."):
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from io import BytesIO
            
            # 2026 verisi hazırla - orijinal format
            data_2026 = full_data[full_data['Year'] == 2026].copy()
            
            # Orijinal Excel'deki kolon yapısına uygun hale getir
            excel_2026 = pd.DataFrame()
            
            for month in range(1, 13):
                month_data = data_2026[data_2026['Month'] == month].copy()
                
                if len(month_data) > 0:
                    # Toplam satırı ekle
                    total_row = pd.DataFrame({
                        'Month': [f'Toplam {month}'],
                        'MainGroup': [None],
                        'Sales': [month_data['Sales'].sum()],
                        'GrossProfit': [month_data['GrossProfit'].sum()],
                        'GrossMargin%': [month_data['GrossProfit'].sum() / month_data['Sales'].sum() if month_data['Sales'].sum() > 0 else 0],
                        'Stock': [month_data['Stock'].mean()],
                        'COGS': [month_data['COGS'].sum()]
                    })
                    
                    month_data_with_total = pd.concat([month_data[['Month', 'MainGroup', 'Sales', 'GrossProfit', 'GrossMargin%', 'Stock', 'COGS']], total_row], ignore_index=True)
                    excel_2026 = pd.concat([excel_2026, month_data_with_total], ignore_index=True)
            
            # Excel dosyasına yaz
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # 2026 sheet'i ekle
                excel_2026.to_excel(writer, sheet_name='2026_Tahmin', index=False)
                
                # Worksheet'i al ve formatla
                workbook = writer.book
                worksheet = workbook['2026_Tahmin']
                
                # Header formatı
                header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                header_font = Font(color="FFFFFF", bold=True)
                
                for cell in worksheet[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                
                # Kolon genişlikleri
                worksheet.column_dimensions['A'].width = 12
                worksheet.column_dimensions['B'].width = 25
                worksheet.column_dimensions['C'].width = 18
                worksheet.column_dimensions['D'].width = 18
                worksheet.column_dimensions['E'].width = 15
                worksheet.column_dimensions['F'].width = 18
                worksheet.column_dimensions['G'].width = 18
                
                # Number formatları
                for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):
                    # Para formatı (Sales, GrossProfit, Stock, COGS)
                    for col_idx in [2, 3, 5, 6]:  # 0-indexed: C, D, F, G
                        if row[col_idx].value and isinstance(row[col_idx].value, (int, float)):
                            row[col_idx].number_format = '#,##0'
                    
                    # Yüzde formatı (GrossMargin%)
                    if row[4].value and isinstance(row[4].value, (int, float)):
                        row[4].number_format = '0.00%'
            
            excel_data = output.getvalue()
            
            st.download_button(
                label="📥 Bütçe Dosyası İndir (Excel)",
                data=excel_data,
                file_name="butce_2026_tahmin.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            st.success("✅ Excel dosyası hazır! İndir butonuna tıklayın.")

# Footer
st.markdown("---")
st.markdown("""
    <div style='text-align: center; color: #666;'>
        <p>2026 Satış Bütçe Tahmin Sistemi | Ay + Ana Grup Bazında Hedefleme</p>
    </div>
""", unsafe_allow_html=True)
